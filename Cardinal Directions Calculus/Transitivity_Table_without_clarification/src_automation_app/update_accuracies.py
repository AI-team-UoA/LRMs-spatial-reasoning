import pandas as pd
from openpyxl import load_workbook

from .calc_jaccard import process_sheet_calc_jaccard

def update_accuracies_Interface():
	"""
	Interface for the function update_accuracies. Collects params from the user and calls the function.
	"""
	excel_path = input("Enter the path to the Excel file with answers: ")
	model_name = input("Enter the model name for which accuracies are being updated: ")
	
	# Call the function to update accuracies
	update_accuracies(excel_path, model_name)


def update_accuracies(excel_path, model_name, review_excel = "all_accuracies.xlsx"):
	"""
	Update the Jaccard Index accuracies in the review Excel file based on the provided Excel file with the answers.
	
	:param excel_path: Path to the Excel file containing answers.
	:param review_excel: Path to the review Excel file where accuracies will be updated.
	:param model_name: Name of the model for which accuracies are being updated.
	"""
	
	# Load the Excel file
	try:
		excel_file = pd.ExcelFile(excel_path)
	except FileNotFoundError:
		print(f"No such file as '{excel_path}'")
		return

	excel_file = pd.ExcelFile(excel_path)
	print(f"Processing file: {excel_path}")
	jaccards_scores = []

	# Iterate through each sheet in the Excel file and calculate the average Jaccard Index
	for idx_sheet_name, sheet_name in enumerate(excel_file.sheet_names):
		print(f"\tProcessing sheet {sheet_name}: {idx_sheet_name}")
		avg_jaccard = process_sheet_calc_jaccard(excel_path, sheet_name, idx_sheet_name)
		jaccards_scores.append(avg_jaccard)

	# Calculate the overall mean Jaccard Index
	if jaccards_scores:
		overall_avg = round(sum(jaccards_scores) / len(jaccards_scores), 2)
		print(f"\n\tOverall mean Jaccard Index: {overall_avg}%")

	    # Record the overall mean Jaccard Index in the all_accuracies.xlsx file
		wb_dest = load_workbook(review_excel)
		ws_dest = wb_dest.active

		found_row = -1
	    # finding the row for the specific model
		for row_idx, row in enumerate(ws_dest.iter_rows(), start=1):
			if row[0].value and str(row[0].value).strip() == model_name:
				found_row = row_idx
				break

			
		if found_row != -1: # If a row for the model name is found, update it
			# record the jacard values for each repetition
			for i, jaccard_score in enumerate(jaccards_scores):
				ws_dest.cell(row=found_row, column=i + 2, value=jaccard_score)

			# Note: The columns with formulas (PI, S, min, max) will be updated
			# when the file is opened in Excel. Openpyxl does not update them automatically
	
		else: # If no row is found, create a new one.
            # get the next empty row
			new_row_idx = ws_dest.max_row + 1
            
            # write the model name
			ws_dest.cell(row=new_row_idx, column=1, value=model_name)
            
            # write Jaccard scores
			for i, jaccard_score in enumerate(jaccards_scores):
				ws_dest.cell(row=new_row_idx, column=i + 2, value=jaccard_score)



	    # save workbook
		wb_dest.save(review_excel)
		print(f"'{review_excel}' updated successfully")