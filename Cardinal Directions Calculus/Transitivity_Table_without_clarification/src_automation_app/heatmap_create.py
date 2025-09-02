import os
from openpyxl import load_workbook, Workbook
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import seaborn as sns

def calculate_jaccard(cell_value):
    if not cell_value or '|' not in cell_value:
        return None

    parts = cell_value.split('|')
    if len(parts) != 2:
        return None

    set1 = {v.strip() for v in parts[0].split(',') if v.strip()}
    set2 = {v.strip() for v in parts[1].split(',') if v.strip()}

    union = set1 | set2
    if not union:
        return 1.0  # Perfect match if both are empty

    intersection = set1 & set2
    return round(len(intersection) / len(union), 4)

def average_jaccard_across_sheets(file_path, sheet_names, output_path):
    # Load workbook and sheets
    wb = load_workbook(file_path, data_only=True)
    sheets = [wb[name] for name in sheet_names]

    # Load labels for headers from the correct file
    labels_file_path = "./correct_result_transitivity_table.xlsx"
    labels_df_full = pd.read_excel(labels_file_path, sheet_name='Φύλλο3', header=None)
    x_labels = labels_df_full.iloc[0, 1:].tolist()
    y_labels = labels_df_full.iloc[1:, 0].tolist()

    max_rows = max(sheet.max_row for sheet in sheets)
    max_cols = max(sheet.max_column for sheet in sheets)

    # Prepare data for DataFrame
    data_list = []
    for row in range(1, max_rows + 1):
        row_data = []
        for col in range(1, max_cols + 1):
            jaccards = []
            for sheet in sheets:
                cell = sheet.cell(row=row, column=col)
                jaccard = calculate_jaccard(cell.value)
                if jaccard is not None:
                    jaccards.append(jaccard)
            if jaccards:
                avg_jaccard = round(sum(jaccards) / len(jaccards), 4)
                row_data.append(avg_jaccard)
        if row_data:
            data_list.append(row_data)

    # Create the DataFrame with the appropriate labels
    df_with_headers = pd.DataFrame(data_list)
    df_with_headers.index = y_labels
    df_with_headers.columns = x_labels
    
    # Create the DataFrame without headers for the heatmap plotting
    df_no_headers = pd.DataFrame(data_list)

    # Save to Excel with two sheets
    output_excel_file = os.path.join(output_path, "average_jaccard_across_all_sheets.xlsx")
    with pd.ExcelWriter(output_excel_file) as writer:
        df_with_headers.to_excel(writer, sheet_name='Averaged Jaccard with Headers', index=True)
        df_no_headers.to_excel(writer, sheet_name='Average Jaccard', index=False, header=False)

    print(f"Averaged Jaccard values written to '{output_excel_file}'")
    return output_excel_file
    

def create_heatmap_Interface():
    """
    Interface for the function create_heatmap. Collects params from the user and calls the function.
    """
    model_name = input("Enter the model name for which the heatmap is being created: ")
    excel_path = input("Enter the path to the Excel file with formatted results: ")
    output_image_path = input("Enter the path where the heatmap image will be saved: ")
    output_excel_path = input("Enter the path where the averaged Jaccard Index values will be saved: ")

	# Call the function to create heatmap
    create_heatmap(model_name, excel_path, output_image_path, output_excel_path)


def create_heatmap(model_name, excel_path, output_image_path, output_excel_path):
    """
    Create a heatmap from the Jaccard Index values in the specified Excel file.
        
    :param excel_path: Path to the Excel file containing Jaccard Index values.
    :param output_image_path: Path where the heatmap image will be saved.
    :param output_excel_path: Path where the averaged Jaccard Index values will be saved.
    """
    matplotlib.use('Agg')

    labels_file_path = "./correct_result_transitivity_table.xlsx" #path to correct tables to take the labels

    # Shared axis labels and annotations
    labels_df_full = pd.read_excel(labels_file_path, sheet_name='Φύλλο3', header=None)
    x_labels = labels_df_full.iloc[0, 1:].tolist()
    y_labels = labels_df_full.iloc[1:, 0].tolist()
    labels_df = labels_df_full.iloc[1:, 1:]
    labels_df.index = y_labels
    labels_df.columns = x_labels
        
    # Define the sheet names to process
    sheet_names = ["Φύλλο1", "Φύλλο2", "Φύλλο3"]  # Replace with your actual sheet names
        
    # Run the function
    jaccard_excel = average_jaccard_across_sheets(excel_path, sheet_names, output_excel_path)

    df = pd.read_excel(jaccard_excel, sheet_name="Averaged Jaccard with Headers", index_col=0)
    df.index = y_labels
    df.columns = x_labels
    df = df.apply(pd.to_numeric, errors='coerce')

    if df.empty:
        print(f"Error: The DataFrame read from '{jaccard_excel}' is empty. No data to plot.")
        return
    
    plt.figure(figsize=(12, 8))
    sns.heatmap(df, annot=labels_df, fmt="", cmap="viridis", cbar=True, linewidths=0.2, linecolor='grey', vmin=0, vmax=1, annot_kws={"size": 7})
    os.makedirs(output_image_path, exist_ok=True)
    plt.title(f"{model_name}")
    plt.savefig(output_image_path + "/heatmap_" + model_name + ".png")
    print(f"Heatmap for {model_name} saved in '{output_image_path}'")