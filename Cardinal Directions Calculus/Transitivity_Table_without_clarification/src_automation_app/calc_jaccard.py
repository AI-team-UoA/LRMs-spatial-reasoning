import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font

def highlight_matches_and_jaccard(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # New sheet name: "Jaccard - SheetName"
    new_sheet_name = f'Jaccard - {sheet_name}'
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    new_sheet = wb.create_sheet(title=new_sheet_name)

    new_row_index = 1
    all_jaccards = []  # Store per-cell Jaccard indices

    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                parts = cell.value.split('|')
                if len(parts) != 2:
                    continue

                values_table1 = {v.strip() for v in parts[0].split(',') if v.strip()}
                values_table2 = {v.strip() for v in parts[1].split(',') if v.strip()}

                black_matches = values_table1 & values_table2
                red_misses = values_table1 - values_table2
                blue_extras = values_table2 - values_table1

                count_black = len(black_matches)
                count_red = len(red_misses)
                count_blue = len(blue_extras)

                denom = count_black + count_red + count_blue
                jaccard = round((count_black * 100 / denom), 2) if denom > 0 else 100.0
                all_jaccards.append(jaccard)

                # Write cell-by-cell values with colors
                new_col_index = 1
                for val in values_table1:
                    color = "000000" if val in values_table2 else "FF0000"
                    new_cell = new_sheet.cell(row=new_row_index, column=new_col_index, value=val)
                    new_cell.font = Font(color=color)
                    new_col_index += 1

                for val in blue_extras:
                    new_cell = new_sheet.cell(row=new_row_index, column=new_col_index, value=val)
                    new_cell.font = Font(color="0000FF")
                    new_col_index += 1

                # Write Jaccard value at end of row
                new_cell = new_sheet.cell(row=new_row_index, column=new_col_index, value=f"Jaccard: {jaccard}%")
                new_cell.font = Font(color="008000", bold=True)
                new_row_index += 1

    # Write average Jaccard at the bottom
    if all_jaccards:
        avg_jaccard = round(sum(all_jaccards) / len(all_jaccards), 2)
        summary_cell = new_sheet.cell(row=new_row_index + 2, column=1, value=f"Average Jaccard Index: {avg_jaccard}%")
        summary_cell.font = Font(color="000080", bold=True)
    else:
        avg_jaccard = 0.0

    wb.save(file_path)
    print(f"Sheet '{sheet_name}' processed. Avg Jaccard: {avg_jaccard}%")
    return avg_jaccard


def process_sheet_calc_jaccard(file_path, sheet_name, i):
    os.makedirs('./temp_files', exist_ok=True)
    new_file_path = "./temp_files/jaccard_results_" + str(i) + ".xlsx"
    shutil.copy(file_path, new_file_path)
    acc = highlight_matches_and_jaccard(new_file_path, sheet_name)
    return(acc)
    