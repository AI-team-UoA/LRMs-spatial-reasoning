import os
import argparse
import time
import json
from dotenv import load_dotenv
from openai import OpenAI
from google import genai
import anthropic
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Get current date
today = datetime.today().strftime('%Y-%m-%d')  # Format: YYYY-MM-DD

#Number of iterarions to do
i_tot = 3

# Load environment variables
load_dotenv()

# Model parameters
models_dict = {
  'o1': {
    'alias': 'o1-2024-12-17', 
    'api-interface': 'openai', 
    'api-key': os.environ.get('OPENAI_API_KEY', 'api-key')
  },
  'o3': {
    'alias': 'o3-2025-04-16', 
    'api-interface': 'openai', 
    'api-key': os.environ.get('OPENAI_API_KEY', 'api-key')
  },
  'o4-mini': {
    'alias': 'o4-mini-2025-04-16', 
    'api-interface': 'openai', 
    'api-key': os.environ.get('OPENAI_API_KEY', 'api-key')
  },
  'deepseek-r1': {
    'alias': 'deepseek-reasoner', 
    'api-interface': 'openai',
    'api-key': os.environ.get('DEEPSEEK_API_KEY', 'api-key'),
    'base-url': os.environ.get('DEEPSEEK_BASE_URL', 'https://api.deepseek.com')
  },
  'grok-3-mini': {
    'alias': 'grok-3-mini',
    'api-interface': 'openai',
    'provider': 'xAI',
    'api-key': os.environ.get('GROK_API_KEY', 'api-key'),
    'base-url': os.environ.get('DEEPSEEK_BASE_URL', 'https://api.x.ai/v1')
  },
  'claude-3.7-sonnet': {
    'alias': 'claude-3-7-sonnet-20250219',
    'api-interface': 'anthropic',
    'api-key': os.environ.get('CLAUDE_API_KEY', 'api-key')
  },
  'gemini-2.5-flash': {
    'alias': 'gemini-2.5-flash-preview-04-17',
    'api-interface': 'genai',
    'api-key': os.environ.get('GEMINI_API_KEY', 'api-key')
  }
}

# Initialize arguments parser
args_parser = argparse.ArgumentParser(description="Perform experiments with different models and arguments")
args_parser.add_argument('--model', choices=['o1', 'o3', 'o4-mini', 'deepseek-r1', 'grok-3-mini','claude-3.7-sonnet','gemini-2.5-flash'], required=True, help='Model to use for experiment.')
args_parser.add_argument('--effort', choices=['low', 'medium', 'high'], default='medium', help='The reasoning effort to use. The higher the effort the higher the computation time and cost.')
args_parser.add_argument('--experiment', choices=['composition', 'composition-disquised'], default='composition', help='The experiment to perform.')
args_parser.add_argument('--relationships_a', default='N,W,E,S,NW,SW,SE,NE,B', help='The first set of relationships to use split with comma(,).')
args_parser.add_argument('--relationships_b', default='N,W,E,S,NW,SW,SE,NE,B', help='The second set of relationships to use split with comma(,).')
args_parser.add_argument('--seed', type=int, default=None, required=False, help='The seed to use for the seed parameter in llm api.')
args_parser.add_argument('--output_dir', default=None, help='The directory to store the results and the outputs of each experiment.')
args = args_parser.parse_args()

for i in range(i_tot):
  
  # Create output directories
  if args.output_dir is None:
    output_dir = args.model+'_def_'+today+'_output_'+str(i+1)
  else:
    output_dir = args.output_dir

  if not os.path.exists(output_dir):
    os.mkdir(output_dir) # main directory

  if not os.path.exists(os.path.join(output_dir, 'json')):
    os.mkdir(os.path.join(output_dir, 'json')) # json outputs directory

  if not os.path.exists(os.path.join(output_dir, 'log')):
    os.mkdir(os.path.join(output_dir, 'log')) # log output directory 

  # Create a new workbook for colour comparison in python
  wb = Workbook()
  #wb.create_sheet(title='Transitivity')
  sheet=wb['Sheet']

  # Initialize llm api interface
  api_key = models_dict[args.model]['api-key']
  if args.model == "gemini-2.5-flash":
    if 'base-url' in models_dict[args.model]:
      client = genai.Client(api_key=api_key, base_url=models_dict[args.model]['base-url'])
    else:
      client = genai.Client(api_key=api_key)
  elif args.model == "claude-3.7-sonnet":
    if 'base-url' in models_dict[args.model]:
      client = anthropic.Anthropic(api_key=api_key, base_url=models_dict[args.model]['base-url'])
    else:
      client = anthropic.Anthropic(api_key=api_key)
  else:
    if 'base-url' in models_dict[args.model]:
      client = OpenAI(api_key=api_key, base_url=models_dict[args.model]['base-url'])
    else:
      client = OpenAI(api_key=api_key)

  # Start the experiments
  model_alias = models_dict[args.model]['alias']

  relationships_a = args.relationships_a.split(',')
  relationships_b = args.relationships_b.split(',')

  system_prompt = "Given a region a, the greatest lower bound or infimum of the projection of a on the x-axis (resp. on the y-axis) is denoted by infx(a) (resp. infy(a)). The least upper bound or the supremum of the projection of a on the x-axis (resp. on the y-axis) is denoted by supx(a) (resp. supy(a)). These bounds define the minimum bounding box of a region a, which is the box formed by the straight lines x=infx(a), x=supx(a), y=infy(a) and y=supy(a). Let us now consider regions that are homeomorphic to the closed unit disk {(x, y): x^2 + y^2 <= 1}. The set of these regions will be denoted by REG. Regions in REG are closed, connected and have connected boundaries. A cardinal direction relation between regions in REG is one of the following relations: B (bounding box), S (South), SW (South West), W (West), NW (North West), N (North), NE (North East), E (East) and SE (South East). These relations are defined as follows: a B b if and only if infx(b) <= infx(a), supx(a) <= supx(b), infy(b) <= infy(a) and supy(a) <= supy(b). a S b if and only if supy(a) <= infy(b), infx(b) <= infx(a) and supx(a) <= supx(b). a SW b if and only if supx(a) <= infx(b) and supy(a) <= infy(b). a W b if and only if supx(a) <= infx(b), infy(b) <= infy(a) and supy(a) <= supy(b). a NW b if and only if supx(a) <= infx(b) and supy(b) <= infy(a). a N b if and only if supy(b) <= infy(a), infx(b) <= infx(a) and supx(a) <= supx(b). a NE b if and only if supx(b) <= infx(a) and supy(b) <= infy(a). a E b if and only if supx(b) <= infx(a), infy(b) <= infy(a) and supy(a) <= supy(b). a SE b if and only if supx(b) <= infx(a) and supy(a) <= infy(b). You are a helpful assistant. I will now give you a question regarding the cardinal direction relations I defined above. The possible answer can be one or more of N, NE, SE, S, E, NW, W, SW, B. No yapping."
  
  main_file = open(os.path.join(output_dir, 'summary.txt'), 'x')
  main_file.write(f"Using model: {args.model} ({model_alias}) with Default effort, seed and temperature\n")
  main_file.write('===========================================================================================================\n\n')

  total_time = 0
  total_completion_tokens = 0
  total_reasoning_tokens = 0

  all_answers_file = open(os.path.join(output_dir, 'all_answers.txt'), 'x', encoding='utf-8')

  #for excel
  row = 1
  col = 1 

  if args.model == "gemini-2.5-flash":
    print('running for gemini-2.5-flash')
    for r1 in relationships_a:
      for r2 in relationships_b:
        user_prompt = "Let " + r1 + " and " + r2 + " be cardinal direction relations. If region x is " + r1 + " of region y and region y is " + r2 + " of region z, then which could the possible relations between region x and region z be?"
        t = time.time()
        response = client.models.generate_content(
          model=model_alias,
          contents=system_prompt+user_prompt,
          config=genai.types.GenerateContentConfig(
          )
        )
        elapsed_time = time.time() - t
        total_time += elapsed_time
        answer = response.text
        if hasattr(response, 'usage_metadata'):
          reasoning_tokens = response.usage_metadata.prompt_token_count
          completion_tokens = response.usage_metadata.thoughts_token_count
        else:
          reasoning_tokens = 0
          completion_tokens = 0
        total_reasoning_tokens += reasoning_tokens
        total_completion_tokens += completion_tokens
        # Write log (txt) output
        with open(os.path.join(output_dir, 'log', f'{r1}_{r2}.txt'), 'x', encoding='utf-8') as log_file:
          log_file.write(f'PROMPT: {user_prompt}\n\nANSWER:\n\n{answer}\n\n\nELAPSED TIME: {elapsed_time:.2f} seconds\nPROMPT TOKENS: {reasoning_tokens}\nTHOUGHTS TOKENS: {completion_tokens}')
        # Write json output
        json_output = {'prompt': user_prompt, 'answer': answer, 'elapsed-time': elapsed_time, 'prompt-tokens': reasoning_tokens, 'thoughts-tokens': completion_tokens}
        with open(os.path.join(output_dir, 'json', f'{r1}_{r2}.json'), 'x', encoding='utf-8') as json_file:
          json_file.write(json.dumps(json_output))
        # Write all answers
        all_answers_file.write(f'For r1: {r1}, r2: {r2}, answer: {answer}\n')
        sheet.cell(column=col, row=row).value= answer
        col = col + 1
      row = row + 1
      col = 1
  elif args.model == "claude-3.7-sonnet":
    print('running for claude-3.7-sonnet')
    for r1 in relationships_a:
      for r2 in relationships_b:
        user_prompt = "Let " + r1 + " and " + r2 + " be cardinal direction relations. If region x is " + r1 + " of region y and region y is " + r2 + " of region z, then which could the possible relations between region x and region z be?"
        t = time.time()
        response = client.messages.create(
          model=model_alias,
          #reasoning_effort=args.effort,
          max_tokens=16000,
          system = system_prompt,
          messages=[
            {
              "role": "user", 
              "content": user_prompt
            }
          ],
          thinking={
              "type": "enabled",
              "budget_tokens": 15500
          }
          #seed = seed_used,
          #store=False
          #seed=args.seed, # First used = 2200017
        )
        elapsed_time = time.time() - t
        total_time += elapsed_time
        #answer = response.choices[0].message.content
        '''answer = response.content[0].text
        reasoning_tokens = response.usage.completion_tokens_details.to_dict().get('input_tokens', 0) if 'completion_tokens_details' in response.usage.to_dict() else 0
        total_reasoning_tokens += reasoning_tokens
        completion_tokens = response.usage.completion_tokens_details.to_dict().get('output_tokens', 0) if 'completion_tokens_details' in response.usage.to_dict() else 0
        total_completion_tokens += completion_tokens'''
        answer = ""
        thinking_text = ""
        for content_block in response.content:
          if content_block.type == "text":
            answer += content_block.text
          elif content_block.type == "thinking":
            thinking_text += content_block.thinking
        reasoning_tokens = response.usage.output_tokens
        total_reasoning_tokens += reasoning_tokens
        completion_tokens = response.usage.output_tokens
        total_completion_tokens += completion_tokens
        # Write log (txt) output
        with open(os.path.join(output_dir, 'log', f'{r1}_{r2}.txt'), 'x', encoding='utf-8') as log_file:
          log_file.write(f'PROMPT: {user_prompt}\n\nTHINKING:\n\n{thinking_text}\n\nANSWER:\n\n{answer}\n\n\nELAPSED TIME: {elapsed_time:.2f} seconds\nOUTPUT TOKENS: {completion_tokens}')
        # Write json output
        json_output = {'prompt': user_prompt, 'answer': answer, 'elapsed-time': elapsed_time, 'reasoning-tokens': str(reasoning_tokens), 'completion-tokens': str(completion_tokens)}
        with open(os.path.join(output_dir, 'json', f'{r1}_{r2}.json'), 'x', encoding='utf-8') as json_file:
          json_file.write(json.dumps(json_output))
        # Write all answers
        all_answers_file.write(f'For r1: {r1}, r2: {r2}, answer: {answer}\n')
        sheet.cell(column=col, row=row).value= answer
        col = col + 1
      row = row + 1
      col = 1
  else:
    print('running for openai api')
    if args.model == "grok-3-mini":
      for r1 in relationships_a:
        for r2 in relationships_b:
          user_prompt = "Let " + r1 + " and " + r2 + " be cardinal direction relations. If region x is " + r1 + " of region y and region y is " + r2 + " of region z, then which could the possible relations between region x and region z be?"
          t = time.time()
          response = client.chat.completions.create(
            model=model_alias,
            #reasoning_effort=args.effort, #IF MODEL = GROK REMOVE IT, ELSE LEAVE IT AS IS
            messages=[
              {
                "role": "system", 
                "content": system_prompt
              },
              {
                "role": "user", 
                "content": user_prompt
              }
            ],
            store=False
          )
          elapsed_time = time.time() - t
          total_time += elapsed_time
          answer = response.choices[0].message.content
          reasoning_tokens = response.usage.completion_tokens_details.to_dict().get('reasoning_tokens', 0) if 'completion_tokens_details' in response.usage.to_dict() else 0
          total_reasoning_tokens += reasoning_tokens
          completion_tokens = response.usage.completion_tokens
          total_completion_tokens += completion_tokens
          # Write log (txt) output
          with open(os.path.join(output_dir, 'log', f'{r1}_{r2}.txt'), 'x', encoding='utf-8') as log_file:
            log_file.write(f'PROMPT: {user_prompt}\n\nANSWER:\n\n{answer}\n\n\nELAPSED TIME: {elapsed_time:.2f} seconds\nREASONING TOKENS: {reasoning_tokens}\nCOMPLETION TOKENS: {completion_tokens}')
          # Write json output
          json_output = {'prompt': user_prompt, 'answer': answer, 'elapsed-time': elapsed_time, 'reasoning-tokens': reasoning_tokens, 'completion-tokens': completion_tokens}
          with open(os.path.join(output_dir, 'json', f'{r1}_{r2}.json'), 'x', encoding='utf-8') as json_file:
            json_file.write(json.dumps(json_output))
          # Write all answers
          all_answers_file.write(f'For r1: {r1}, r2: {r2}, answer: {answer}\n')
          sheet.cell(column=col, row=row).value= answer
          col = col + 1
        row = row + 1
        col = 1
    elif args.model == "deepseek-r1":
      for r1 in relationships_a:
        for r2 in relationships_b:
          user_prompt = "Let " + r1 + " and " + r2 + " be cardinal direction relations. If region x is " + r1 + " of region y and region y is " + r2 + " of region z, then which could the possible relations between region x and region z be?"
          t = time.time()
          response = client.chat.completions.create(
            model=model_alias,
            reasoning_effort=args.effort, #IF MODEL = GROK REMOVE IT, ELSE LEAVE IT AS IS
            messages=[
              {
                "role": "system", 
                "content": system_prompt
              },
              {
                "role": "user", 
                "content": user_prompt
              }
            ],
            store=False
          )
          elapsed_time = time.time() - t
          total_time += elapsed_time
          #answer = response.choices[0].message.content
          answer = response.choices[0].message.content
          reasoning_tokens = response.usage.completion_tokens_details.to_dict().get('reasoning_tokens', 0) if 'completion_tokens_details' in response.usage.to_dict() else 0
          total_reasoning_tokens += reasoning_tokens
          completion_tokens = response.usage.completion_tokens
          total_completion_tokens += completion_tokens
          # Write log (txt) output
          with open(os.path.join(output_dir, 'log', f'{r1}_{r2}.txt'), 'x', encoding='utf-8') as log_file:
            log_file.write(f'PROMPT: {user_prompt}\n\nANSWER:\n\n{answer}\n\n\nELAPSED TIME: {elapsed_time:.2f} seconds\nREASONING TOKENS: {reasoning_tokens}\nCOMPLETION TOKENS: {completion_tokens}')
          # Write json output
          json_output = {'prompt': user_prompt, 'answer': answer, 'elapsed-time': elapsed_time, 'reasoning-tokens': reasoning_tokens, 'completion-tokens': completion_tokens}
          with open(os.path.join(output_dir, 'json', f'{r1}_{r2}.json'), 'x', encoding='utf-8') as json_file:
            json_file.write(json.dumps(json_output))
          # Write all answers
          all_answers_file.write(f'For r1: {r1}, r2: {r2}, answer: {answer}\n')
          sheet.cell(column=col, row=row).value= answer
          col = col + 1
        row = row + 1
        col = 1
    else:
      for r1 in relationships_a:
        for r2 in relationships_b:
          user_prompt = "Let " + r1 + " and " + r2 + " be cardinal direction relations. If region x is " + r1 + " of region y and region y is " + r2 + " of region z, then which could the possible relations between region x and region z be?"
          t = time.time()
          response = client.chat.completions.create(
            model=model_alias,
            reasoning_effort=args.effort, #IF MODEL = GROK REMOVE IT, ELSE LEAVE IT AS IS
            messages=[
              {
                "role": "system", 
                "content": system_prompt
              },
              {
                "role": "user", 
                "content": user_prompt
              }
            ],
            store=False
          )
          elapsed_time = time.time() - t
          total_time += elapsed_time
          #answer = response.choices[0].message.content
          answer = response.choices[0].message.content
          reasoning_tokens = response.usage.completion_tokens_details.to_dict().get('reasoning_tokens', 0) if 'completion_tokens_details' in response.usage.to_dict() else 0
          total_reasoning_tokens += reasoning_tokens
          completion_tokens = response.usage.completion_tokens
          total_completion_tokens += completion_tokens
          # Write log (txt) output
          with open(os.path.join(output_dir, 'log', f'{r1}_{r2}.txt'), 'x', encoding='utf-8') as log_file:
            log_file.write(f'PROMPT: {user_prompt}\n\nANSWER:\n\n{answer}\n\n\nELAPSED TIME: {elapsed_time:.2f} seconds\nREASONING TOKENS: {reasoning_tokens}\nCOMPLETION TOKENS: {completion_tokens}')
          # Write json output
          json_output = {'prompt': user_prompt, 'answer': answer, 'elapsed-time': elapsed_time, 'reasoning-tokens': reasoning_tokens, 'completion-tokens': completion_tokens}
          with open(os.path.join(output_dir, 'json', f'{r1}_{r2}.json'), 'x', encoding='utf-8') as json_file:
            json_file.write(json.dumps(json_output))
          # Write all answers
          all_answers_file.write(f'For r1: {r1}, r2: {r2}, answer: {answer}\n')
          sheet.cell(column=col, row=row).value= answer
          col = col + 1
        row = row + 1
        col = 1
  # Write in summary file
  main_file.write(f'TOTAL TIME: {total_time:.2f}\nAVERAGE TIME: {total_time/(len(relationships_a) * len(relationships_b)):.2f}\n\nTOTAL REASONING TOKENS: {total_reasoning_tokens}\nAVERAGE REASONING TOKENS: {int(total_reasoning_tokens/(len(relationships_a) * len(relationships_b)))}\nTOTAL COMPLETION TOKENS: {total_completion_tokens}\nAVERAGE COMPLETION TOKENS: {int(total_completion_tokens/(len(relationships_a) * len(relationships_b)))}\n')
  all_answers_file.close()
  main_file.close()
  wb.save(output_dir+'.xlsx')
